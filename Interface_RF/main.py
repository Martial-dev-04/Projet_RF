from flask import Flask, render_template, Response, request, send_file
import tempfile
import cv2
import os
import face_recognition
from deepface import DeepFace
import numpy as np
import joblib
import threading
import hashlib
from time import sleep
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime




app = Flask(__name__)

mois_fr = {
    "January": "janvier", "February": "février", "March": "mars",
    "April": "avril", "May": "mai", "June": "juin",
    "July": "juillet", "August": "août", "September": "septembre",
    "October": "octobre", "November": "novembre", "December": "décembre"
}

jours_fr = {
    "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
    "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi",
    "Sunday": "Dimanche"
}

# Charger le modèle
model_data = joblib.load("C:/PROJETS/Reconnaissance_faciale/Projet_RF/face_model.pkl")
svm_model = model_data["model"]
label_encoder = model_data["encoder"]
deepface_model = model_data["deepface_model"]

fichier_excel = "C:/PROJETS/Reconnaissance_faciale/Projet_RF/Liste_de_présence.xlsx"
try:
    if os.path.exists(fichier_excel):
        df = load_workbook(fichier_excel)
    else:
        df = Workbook()
        df.remove(df.active)
    date_aujourdhui = datetime.now().strftime("%d-%m-%Y")
    if date_aujourdhui not in df.sheetnames:
        feuille = df.create_sheet(date_aujourdhui)
        feuille.append(["Noms et Prénoms", "Heures", "Dates"])
    else:
        feuille = df[date_aujourdhui]

    personnes_deja_presentes = set([ligne[0] for ligne in feuille.iter_rows(min_row=2, values_only=True) if ligne[0]])
    visages_reconnus = {}  # cache des visages reconnus


    def gen():
        cap = cv2.VideoCapture(0)
        sleep(2)  # Laisse le temps à la webcam de se stabiliser

        while True:
            success, frame = cap.read()
            if not success:
                break

            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            face_locations = face_recognition.face_locations(rgb_small_frame)

            for top, right, bottom, left in face_locations:
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4

                face_img = frame[top:bottom, left:right]
                face_id = hashlib.md5(face_img.tobytes()).hexdigest()

                if face_id in visages_reconnus:
                    name = visages_reconnus[face_id]["name"]
                    confidence = visages_reconnus[face_id]["confidence"]

                if face_img.size == 0 or face_img.shape[0] == 0 or face_img.shape[1] == 0:
                    continue
                key = f"{top}-{right}-{bottom}-{left}"
                if key in visages_reconnus:
                    name, confidence = visages_reconnus[key]

                else:
                    try:
                        embedding = DeepFace.represent(
                            img_path=face_img,
                            model_name=deepface_model,
                            enforce_detection=False
                        )[0]["embedding"]

                        proba = svm_model.predict_proba([embedding])[0]
                        pred_label = np.argmax(proba)
                        confidence = np.max(proba)

                        if confidence >= 0.90:
                            name = label_encoder.inverse_transform([pred_label])[0]
                            visages_reconnus[key] = (name, confidence)

                            if name not in personnes_deja_presentes:
                                maintenant = datetime.now()
                                heure = maintenant.strftime("%H:%M:%S")
                                date = maintenant.strftime("%d/%m/%Y")
                                feuille.append([name, heure, date])
                                df.save(fichier_excel)
                                personnes_deja_presentes.add(name)
                                print(f"✅ Présence enregistrée : {name} à {heure} le {date}")
                        else:
                            name = "Inconnu"
                    except Exception:
                        name = "Inconnu"
                        confidence = 0

                # Affichage
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0) if name != "Inconnu" else (0, 0, 255), 2)
                cv2.putText(
                    frame, f"{name} ({confidence:.2f})",
                    (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                    (0, 255, 0) if name != "Inconnu" else (0, 0, 255), 2
                )

            ret, buffer = cv2.imencode('.jpg', frame)
            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
except Exception as e:
    print("Erreur pendant l'enregistrement: ", e)

finally:
    df.close()
    print("\n📁 Session terminée, fichier sauvegardé !")


def appliquer_mise_en_forme(feuille):
    # Vérifier que le 1er ligne correspond bien à un en-tête
    if feuille.max_row >= 1:
        # Appliquer style à la 1er ligne (en-tête)
        for col in range(1, feuille.max_column + 1):
            cell = feuille.cell(row=1, column=col)
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ajuster automatiquement la largeur des colones
        for col in range(1, feuille.max_column + 1):
                    lettre = get_column_letter(col)
                    max_length = 0
                    for row in feuille.iter_rows(min_row=1, max_col=col, max_row=feuille.max_row):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    feuille.column_dimensions[lettre].width = max_length + 2
@app.route('/')
def acceuil():
    return render_template('base.html')


@app.route('/camera')
def cam():
    return render_template('camera.html')


@app.route('/video')
def video():
    return Response(gen(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/presence')
def presence():
    chemin_excel = "C:/PROJETS/Reconnaissance_faciale/Projet_RF/Liste_de_présence.xlsx"

    if not os.path.exists(chemin_excel):
        return render_template("presence.html", donnees = [], feuilles_disponibles=[])
    
    
    try:

        wb = load_workbook(chemin_excel)
        feuilles_disponibles = wb.sheetnames
        #print("Feuilles disponibles: ",feuilles_disponibles)

        feuilles_formatees = []
        feuille_org = {}
        # Appliquer le mise en forme à chaque feuille
        for nom in wb.sheetnames:
            feuille = wb[nom]
            # Appliquer style à la 1er ligne (en-tête)
            for col in range(1, feuille.max_column + 1):
                cell = feuille.cell(row=1, column=col)
                cell.font = Font(name='Cambria', size=11, bold=True)
                cell.fill = PatternFill(start_color="BDD4EE", end_color="BDD4EE", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajuster automatiquement la largeur des colones
            for col in range(1, feuille.max_column + 1):
                lettre = get_column_letter(col)
                max_length = 0
                for row in feuille.iter_rows(min_row=1, max_col=col, max_row=feuille.max_row):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                feuille.column_dimensions[lettre].width = max_length + 2
            # Enregistrer les modifications
            wb.save(chemin_excel)

            # Essayer de convertir avec plusieurs formats
            try:
                
                date_obj_temp = datetime.strptime(nom, "%d-%m-%Y")
                formatee = date_obj_temp.strftime("%d-%m-%Y")
                feuilles_formatees.append(formatee)
                feuille_org[formatee] = nom
            except Exception as e:
                print(f"Feuille ignorée ({nom}) : {e}")
                continue
        #print("Feuilles disponibles: ", feuilles_formatees)
        feuille_choisie = request.args.get('feuille')

        nom_feuille_aujoudhui = datetime.now().strftime("%d-%m-%Y")

        if not feuille_choisie:
            feuille_choisie = nom_feuille_aujoudhui

        nom_org = feuille_org.get(feuille_choisie)
        donnees = []
        total = 0
        date_affichage = "Non disponible"

        if nom_org and nom_org in wb.sheetnames:
            try:
                date_obj = datetime.strptime(nom_org, "%d-%m-%Y")
                jour = jours_fr[date_obj.strftime("%A")]
                mois = mois_fr[date_obj.strftime("%B")]
                date_affichage = f" {jour} {date_obj.day:02d} {mois} {date_obj.year} "
            except Exception as e:
                print(f"Erreur de format de date : {e}")
                date_affichage = nom_org  # fallback si erreur

            feuille = wb[nom_org]
            for row in feuille.iter_rows(min_row=2, values_only=True):
                donnees.append(row)
                total += 1
            #print("Total des présences: ", total)

    except FileNotFoundError:
        print("fichier Excel introuvable")


    return render_template('presence.html', 
                           donnees=donnees, 
                           feuilles_disponibles=feuilles_formatees,
                           request=request,
                           nom_feuille=date_affichage,
                           total=total)

@app.route('/telecharger_feuille')
def telecharger_feuille():
    from openpyxl import Workbook
    import tempfile

    feuille_selectionnee = request.args.get('feuille')
    chemin_fichier = "C:/PROJETS/Reconnaissance_faciale/Projet_RF/Liste_de_présence.xlsx"

    if not os.path.exists(chemin_fichier):
        return "❌ Fichier Excel introuvable", 404

    wb = load_workbook(chemin_fichier)
    
    if not feuille_selectionnee:
        feuille_selectionnee = datetime.now().strftime("%d-%m-%Y")

    # 🔁 Rechercher le vrai nom de la feuille contenant la date choisie
    feuille_exacte = None
    for nom in wb.sheetnames:
        try:
            date_obj = datetime.strptime(nom, "%d-%m-%Y")
            if date_obj.strftime("%d-%m-%Y") == feuille_selectionnee:
                feuille_exacte = nom
                break
        except:
            continue

    if not feuille_exacte:
        return f"❌ Feuille correspondant à la date {feuille_selectionnee} introuvable.", 404

    # 📝 Créer un nouveau classeur temporaire avec uniquement cette feuille
    wb_temp = Workbook()
    wb_temp.remove(wb_temp.active)
    feuille_source = wb[feuille_exacte]
    feuille_nouvelle = wb_temp.create_sheet(title=feuille_exacte)

    for ligne in feuille_source.iter_rows(values_only=True):
        feuille_nouvelle.append(ligne)

    # 💾 Sauvegarde dans un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb_temp.save(temp_file.name)

    return send_file(temp_file.name, as_attachment=True, download_name=f"Présence_{feuille_selectionnee}.xlsx")


@app.route('/telecharger_excel')
def telecharger_excel():
    chemin_excel = "C:/PROJETS/Reconnaissance_faciale/Projet_RF/Liste_de_présence.xlsx"
    if os.path.exists(chemin_excel):
        return send_file(chemin_excel, as_attachment=True, download_name="Liste_de_presence_complet.xlsx")
    return "Fichier introuvable", 404



if __name__ == '__main__':
    #threading.Thread(target=face_recognition_thread, daemon=True).start()
    app.run(host='0.0.0.0', port=5000, debug=True)